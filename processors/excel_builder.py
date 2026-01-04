"""
Excel Builder for Master Workbook (STEP 6)
- Creates Master Sheet with all complaints
- Adds Crime-type-wise sheets (grouped by Type_of_Cybercrime)
- Adds Platform-wise sheets (grouped by Platform_Involved)
- Adds Possible Duplicates sheet (flag only; no merge)

Beginner-friendly comments: explains pandas + openpyxl usage.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

REQUIRED_COLUMNS = [
    'Complaint_ID',
    'Complaint_Date_Time',
    'Complainant_Name',
    'Mobile_Number',
    'Email',
    'District',
    'Police_Station',
    'Type_of_Cybercrime',
    'Platform_Involved',
    'Amount_Lost',
    'Current_Status',
]


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure all required columns exist, filling blanks as needed."""
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            df[col] = ''
    return df[REQUIRED_COLUMNS]


def _grouped_sheet(wb, name: str, df: pd.DataFrame, by: str):
    """
    Create a grouped sheet where each group's rows are written sequentially.
    Example: Crime-type-wise or Platform-wise.
    """
    ws = wb.create_sheet(title=name)
    # Write a header row
    ws.append(REQUIRED_COLUMNS)
    # Sort by grouping column for readability
    df_sorted = df.sort_values(by=[by, 'Complaint_ID'])
    for row in dataframe_to_rows(df_sorted[REQUIRED_COLUMNS], index=False, header=False):
        ws.append(row)


def _duplicates_sheet(wb, df: pd.DataFrame):
    """
    Build a sheet listing possible duplicates.
    Rule: same Complaint_ID or same Mobile_Number appear more than once.
    """
    ws = wb.create_sheet(title='Possible_Duplicates')
    ws.append(REQUIRED_COLUMNS + ['Duplicate_Reason'])

    # Find duplicates by Complaint_ID or Mobile_Number
    dup_id = df[df['Complaint_ID'].astype(str).str.len() > 0]
    dup_id = dup_id[dup_id['Complaint_ID'].duplicated(keep=False)]
    dup_id['Duplicate_Reason'] = 'Duplicate Complaint_ID'

    dup_mobile = df[df['Mobile_Number'].astype(str).str.len() > 0]
    dup_mobile = dup_mobile[dup_mobile['Mobile_Number'].duplicated(keep=False)]
    dup_mobile['Duplicate_Reason'] = 'Duplicate Mobile_Number'

    dup_df = pd.concat([dup_id, dup_mobile], ignore_index=True)
    if not dup_df.empty:
        for row in dataframe_to_rows(dup_df[REQUIRED_COLUMNS + ['Duplicate_Reason']], index=False, header=False):
            ws.append(row)


def build_master_workbook(complaints: list, output_path: str):
    """
    STEP 6: After all files are processed, build the Excel workbook.
    - Master Sheet: all complaints
    - Crime-type-wise sheets
    - Platform-wise sheets
    - Possible duplicates sheet
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Convert complaints list (list of dicts) to DataFrame
    df = pd.DataFrame(complaints)
    df = _ensure_columns(df)

    # Write initial workbook with pandas
    df.to_excel(output_path, index=False, sheet_name='Master')

    # Load with openpyxl to add more sheets
    wb = load_workbook(output_path)

    # Crime-type-wise sheet
    _grouped_sheet(wb, 'By_Crime_Type', df, 'Type_of_Cybercrime')

    # Platform-wise sheet
    _grouped_sheet(wb, 'By_Platform', df, 'Platform_Involved')

    # Possible duplicates sheet
    _duplicates_sheet(wb, df)

    # Save workbook
    wb.save(output_path)
